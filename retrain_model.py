from __future__ import annotations

import argparse
import json
from pathlib import Path

import joblib
import numpy as np
import openpyxl
import pandas as pd
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, brier_score_loss, log_loss, roc_auc_score


FEATURES = [
    "Elo diff", "Record diff", "Recent diff", "Strike diff", "TD diff",
    "Control diff", "Reach diff", "Age advantage",
]
FIGHTER_COLUMNS = [
    "Fighter", "UFC fights", "Wins", "Losses", "Draws", "Elo",
    "Smoothed win %", "Recent 5", "Adj strike diff/min", "Adj TD diff/15",
    "Adj control min/15", "Reach", "Age", "Last fight",
]


def load_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
    worksheet = workbook[sheet_name]
    headers = [worksheet.cell(5, column).value for column in range(1, worksheet.max_column + 1)]
    rows = list(worksheet.iter_rows(min_row=6, values_only=True))
    return pd.DataFrame(rows, columns=headers)


def symmetric_probability(model, matrix):
    direct = model.predict_proba(matrix)[:, 1]
    reverse = 1.0 - model.predict_proba(-matrix)[:, 1]
    return np.clip((direct + reverse) / 2.0, 1e-6, 1 - 1e-6)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, required=True)
    parser.add_argument("--output", type=Path, default=Path(__file__).resolve().parent / "data")
    args = parser.parse_args()
    args.output.mkdir(parents=True, exist_ok=True)

    history = load_sheet(args.source, "FIGHT HISTORY")
    history["Date"] = pd.to_datetime(history["Date"], errors="coerce")
    history = history[history["A result"].isin([0, 1])].copy()
    history[FEATURES] = history[FEATURES].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    matrix = history[FEATURES].to_numpy(dtype=float)
    result = history["A result"].astype(int).to_numpy()
    training = history["Date"] < pd.Timestamp("2023-01-01")
    calibration = (history["Date"] >= pd.Timestamp("2023-01-01")) & (history["Date"] < pd.Timestamp("2025-01-01"))
    holdout = (history["Date"] >= pd.Timestamp("2025-01-01")) & (history["Date"] <= pd.Timestamp("2026-08-08"))
    params = dict(
        n_estimators=200, learning_rate=0.04, max_depth=2,
        min_samples_leaf=30, subsample=0.80, random_state=42,
    )

    validation_model = GradientBoostingClassifier(**params)
    validation_model.fit(
        np.vstack([matrix[training], -matrix[training]]),
        np.concatenate([result[training], 1 - result[training]]),
    )
    raw_calibration = symmetric_probability(validation_model, matrix[calibration])
    calibration_logit = np.log(raw_calibration / (1 - raw_calibration)).reshape(-1, 1)
    calibration_result = result[calibration]
    calibrator = LogisticRegression(fit_intercept=False, C=1_000_000, solver="lbfgs")
    calibrator.fit(
        np.vstack([calibration_logit, -calibration_logit]),
        np.concatenate([calibration_result, 1 - calibration_result]),
    )
    slope = float(calibrator.coef_[0, 0])
    raw_holdout = symmetric_probability(validation_model, matrix[holdout])
    holdout_probability = 1 / (1 + np.exp(-slope * np.log(raw_holdout / (1 - raw_holdout))))
    holdout_result = result[holdout]

    production_model = GradientBoostingClassifier(**params)
    production_model.fit(np.vstack([matrix, -matrix]), np.concatenate([result, 1 - result]))
    metrics = {
        "training_fights": int(len(history)),
        "holdout_fights": int(holdout.sum()),
        "accuracy": float(accuracy_score(holdout_result, holdout_probability >= 0.5)),
        "brier": float(brier_score_loss(holdout_result, holdout_probability)),
        "log_loss": float(log_loss(holdout_result, holdout_probability)),
        "auc": float(roc_auc_score(holdout_result, holdout_probability)),
    }
    bundle = {
        "model": production_model,
        "calibration_slope": slope,
        "features": FEATURES,
        "metrics": metrics,
        "importance": dict(zip(FEATURES, production_model.feature_importances_)),
        "version": "UFC-GB-Streamlit-v1",
    }
    joblib.dump(bundle, args.output / "model_bundle.joblib", compress=3)

    fighters = load_sheet(args.source, "FIGHTER DATA")[FIGHTER_COLUMNS].copy()
    fighters.to_csv(args.output / "fighter_snapshot.csv.gz", index=False, compression="gzip")
    (args.output / "model_card.json").write_text(json.dumps(metrics, indent=2))
    print(json.dumps(metrics, indent=2))


if __name__ == "__main__":
    main()
